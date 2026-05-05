import sys, os, tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../../python"))

from generator import generate
from openpyxl import load_workbook

FIXTURE = os.path.join(os.path.dirname(__file__), "../fixtures/sample_template.xlsx")
IMAGE = os.path.join(os.path.dirname(__file__), "../fixtures/sample_image.png")

def _base_request(output_dir):
    return {
        "template_path": FIXTURE,
        "fields": [
            {"name": "编号", "type": "text", "sheet": "Sheet1", "cell": "B3"},
            {"name": "姓名", "type": "text", "sheet": "Sheet1", "cell": "C5"},
        ],
        "rows": [
            {"编号": "001", "姓名": "张三"},
            {"编号": "002", "姓名": "李四"},
        ],
        "output_dir": output_dir,
        "filename_pattern": "{{编号}}_报告.xlsx",
    }

def test_generates_correct_number_of_files():
    with tempfile.TemporaryDirectory() as tmpdir:
        result = generate(_base_request(tmpdir))
        assert len(result["results"]) == 2

def test_output_files_exist():
    with tempfile.TemporaryDirectory() as tmpdir:
        result = generate(_base_request(tmpdir))
        for r in result["results"]:
            assert r["status"] == "success"
            assert os.path.exists(r["file"])

def test_text_fields_are_filled():
    with tempfile.TemporaryDirectory() as tmpdir:
        result = generate(_base_request(tmpdir))
        wb = load_workbook(result["results"][0]["file"])
        ws = wb["Sheet1"]
        assert ws["B3"].value == "001"
        assert ws["C5"].value == "张三"

def test_filename_pattern_applied():
    with tempfile.TemporaryDirectory() as tmpdir:
        result = generate(_base_request(tmpdir))
        names = [os.path.basename(r["file"]) for r in result["results"]]
        assert "001_报告.xlsx" in names
        assert "002_报告.xlsx" in names

def test_missing_field_is_skipped_not_crashed():
    with tempfile.TemporaryDirectory() as tmpdir:
        req = _base_request(tmpdir)
        req["rows"] = [{"编号": "003"}]  # 缺 姓名
        result = generate(req)
        assert result["results"][0]["status"] == "success"

def test_image_field_is_inserted():
    with tempfile.TemporaryDirectory() as tmpdir:
        req = {
            "template_path": FIXTURE,
            "fields": [
                {"name": "图片1", "type": "image", "sheet": "Sheet1", "cell": "A1"},
            ],
            "rows": [{"图片1": IMAGE}],
            "output_dir": tmpdir,
            "filename_pattern": "img_test.xlsx",
        }
        result = generate(req)
        assert result["results"][0]["status"] == "success"
        wb = load_workbook(result["results"][0]["file"])
        assert len(wb["Sheet1"]._images) == 1
