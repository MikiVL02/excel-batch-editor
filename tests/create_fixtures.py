from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"
ws["B3"] = "{{编号}}"
ws["C5"] = "{{姓名}}"
ws["D7"] = "normal_value"  # 非占位符，不应被识别
ws["E9"] = "{{客户名}}和{{编号}}"  # 同一单元格内有两个占位符
wb.save("tests/fixtures/sample_template.xlsx")
print("fixture created")
