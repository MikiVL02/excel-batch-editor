import re
import os
import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

PLACEHOLDER_RE = re.compile(r"\{\{(.+?)\}\}")

def _apply_filename(pattern: str, row: dict) -> str:
    def replace(m):
        return str(row.get(m.group(1), m.group(0)))
    return PLACEHOLDER_RE.sub(replace, pattern)

def generate(req: dict) -> dict:
    template_path = req["template_path"]
    fields = req["fields"]
    rows = req.get("rows")
    data_file_path = req.get("data_file_path")
    output_dir = req["output_dir"]
    filename_pattern = req["filename_pattern"]

    if data_file_path and rows is None:
        dwb = load_workbook(data_file_path, data_only=True)
        dws = dwb.active
        headers = [cell.value for cell in next(dws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in dws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: v for i, v in enumerate(row) if i < len(headers)})

    os.makedirs(output_dir, exist_ok=True)
    results = []

    for i, row in enumerate(rows):
        filename = _apply_filename(filename_pattern, row)
        output_path = os.path.join(output_dir, filename)
        shutil.copy2(template_path, output_path)

        try:
            wb = load_workbook(output_path)
            for field in fields:
                value = row.get(field["name"])
                if value is None:
                    continue
                ws = wb[field["sheet"]]
                if field["type"] == "text":
                    ws[field["cell"]] = value
                elif field["type"] == "image":
                    if os.path.exists(str(value)):
                        img = XLImage(str(value))
                        ws.add_image(img, field["cell"])
                elif field["type"] == "table_range":
                    start_cell = ws[field["cell"]]
                    start_row = start_cell.row
                    start_col = start_cell.column
                    for r_idx, data_row in enumerate(value):
                        for c_idx, cell_value in enumerate(data_row):
                            ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=cell_value)
            wb.save(output_path)
            results.append({"row": i, "status": "success", "file": output_path})
        except Exception as e:
            results.append({"row": i, "status": "error", "error": str(e), "file": output_path})

    return {"results": results}
