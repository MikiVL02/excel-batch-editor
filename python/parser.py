import re
from openpyxl import load_workbook

PLACEHOLDER_RE = re.compile(r"\{\{(.+?)\}\}")

def parse_template(file_path: str) -> dict:
    wb = load_workbook(file_path, data_only=False)
    sheets = wb.sheetnames
    placeholders = []

    for sheet_name in sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    m = PLACEHOLDER_RE.search(cell.value)
                    if m:
                        placeholders.append({
                            "name": m.group(1),
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                        })

    return {"sheets": sheets, "placeholders": placeholders}
